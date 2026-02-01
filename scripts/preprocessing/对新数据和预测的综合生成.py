import argparse
import csv
import io
import math
import re
from pathlib import Path

import numpy as np
import pandas as pd


def normalize_name(x: object) -> str:
    s = "" if x is None else str(x)
    s = s.replace("\ufeff", "").strip()
    s = s.replace("’", "'").replace("`", "'")
    s = re.sub(r"\s+", " ", s)
    s = s.strip(' "')
    return s


def _weighted_median(values: np.ndarray, weights: np.ndarray) -> float:
    mask = np.isfinite(values) & np.isfinite(weights) & (weights > 0)
    v = values[mask]
    w = weights[mask]
    if v.size == 0:
        return float("nan")
    order = np.argsort(v)
    v = v[order]
    w = w[order]
    cum = np.cumsum(w)
    cutoff = 0.5 * float(cum[-1])
    idx = int(np.searchsorted(cum, cutoff, side="left"))
    return float(v[min(max(idx, 0), v.size - 1)])


def _robust_filter_by_mad(x: np.ndarray, k: float = 5.0) -> np.ndarray:
    x = x[np.isfinite(x)]
    if x.size < 5:
        return x
    med = float(np.median(x))
    mad = float(np.median(np.abs(x - med)))
    if mad <= 0:
        return x
    z = np.abs(x - med) / (1.4826 * mad)
    return x[z <= k]


def _sniff_dialect(text: str) -> csv.Dialect:
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
    except Exception:
        return csv.excel


def load_real_fan_table(real_fan_path: Path) -> pd.DataFrame:
    def parse_rows(rows: list[list[str]]) -> pd.DataFrame:
        season_re = re.compile(r"\((\d+)\)")
        season_re_cn = re.compile(r"（(\d+)）")
        season_re_trailing = re.compile(r"(\d+)\s*$")
        number_re = re.compile(r"[-+]?\d+(?:\.\d+)?")

        current_season = None
        name_idx = None
        fan_idx = None
        evidence_idx = None

        out_rows = []
        for row in rows:
            row = [normalize_name(c) for c in row]
            if not any(row):
                continue

            joined = " ".join([c for c in row if c])
            if "信息表" in joined:
                m = season_re.search(joined) or season_re_cn.search(joined) or season_re_trailing.search(joined)
                if m:
                    current_season = int(m.group(1))
                continue

            header_hit = any(("选手" in c) or ("Celebrity" in c) for c in row) and any(("粉丝" in c) or ("受众" in c) for c in row)
            if header_hit:
                name_idx = None
                fan_idx = None
                evidence_idx = None
                for i, c in enumerate(row):
                    if (("选手" in c) or ("Celebrity" in c)) and name_idx is None:
                        name_idx = i
                    if (("粉丝" in c) or ("受众" in c)) and fan_idx is None:
                        fan_idx = i
                    if ("依据" in c) or ("说明" in c):
                        evidence_idx = i
                continue

            if current_season is None:
                continue

            if name_idx is not None and name_idx < len(row):
                name = row[name_idx]
            else:
                nonempty = [c for c in row if c]
                name = nonempty[0] if nonempty else ""

            if fan_idx is not None and fan_idx < len(row):
                fan_cell = row[fan_idx]
            else:
                fan_cell = ""
                for c in row:
                    if number_re.search(c or ""):
                        fan_cell = c
                        break

            nums = number_re.findall(fan_cell or "")
            if not nums:
                continue
            fan_wan = float(nums[0])

            evidence = ""
            if evidence_idx is not None and evidence_idx < len(row):
                evidence = row[evidence_idx]

            if not name:
                continue

            out_rows.append(
                {
                    "Season": int(current_season),
                    "Name": normalize_name(name),
                    "Real_Fan_Wan": fan_wan,
                    "Evidence": evidence,
                }
            )

        df = pd.DataFrame(out_rows)
        return df

    b = real_fan_path.read_bytes()
    if b[:4] == b"PK\x03\x04":
        try:
            xls_raw = pd.read_excel(io.BytesIO(b), header=None)
            rows = xls_raw.fillna("").astype(str).values.tolist()
        except Exception:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(io.BytesIO(b), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = []
            for r in ws.iter_rows(values_only=True):
                rows.append(["" if v is None else str(v) for v in r])
        df = parse_rows(rows)
    else:
        encodings = ["utf-8-sig", "utf-8", "gb18030", "gbk", "latin1"]
        text = None
        for enc in encodings:
            try:
                text = b.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            raise RuntimeError("Unable to decode real_fan.csv with common encodings")

        dialect = _sniff_dialect(text)
        reader = csv.reader(io.StringIO(text), dialect=dialect)
        df = parse_rows([row for row in reader])

    if df.empty:
        raise RuntimeError("Parsed real_fan table is empty; check formatting/encoding")

    df["name_key"] = df["Name"].map(lambda s: normalize_name(s).lower())
    df["Real_Fan"] = (pd.to_numeric(df["Real_Fan_Wan"], errors="coerce") * 10000.0).round().astype("Int64")
    df = df.dropna(subset=["Real_Fan"]).copy()
    df["Real_Fan"] = df["Real_Fan"].astype("int64")
    df = df.sort_values(["Season", "Name"]).drop_duplicates(subset=["Season", "name_key"], keep="first")
    return df


def load_predictions(pred_path: Path) -> pd.DataFrame:
    df = pd.read_csv(pred_path, encoding="utf-8-sig")
    for c in ["Season", "Week"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")
    df["Fan_Vote_Percent"] = pd.to_numeric(df["Fan_Vote_Percent"], errors="coerce")
    df["Name"] = df["Name"].map(normalize_name)
    df["name_key"] = df["Name"].map(lambda s: normalize_name(s).lower())
    df = df.dropna(subset=["Season", "Week", "Fan_Vote_Percent", "name_key"]).copy()
    df["Season"] = df["Season"].astype("int64")
    df["Week"] = df["Week"].astype("int64")
    return df


def estimate_totals_and_votes(merged: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = merged.copy()
    df = df[(df["Fan_Vote_Percent"] > 0) & (df["Real_Fan"] > 0)].copy()

    week_rows = []
    pred_rows = []
    metrics_rows = []

    beta_grid = np.concatenate(
        [
            np.linspace(0.3, 0.9, 7),
            np.linspace(1.0, 2.5, 16),
            np.linspace(2.6, 4.0, 8),
        ]
    )
    lambda_grid = np.linspace(0.05, 1.0, 20)

    for (season, week), g in df.groupby(["Season", "Week"], sort=True):
        g = g.copy()
        g["implied_total"] = g["Real_Fan"].astype(float) / g["Fan_Vote_Percent"].astype(float)
        p = g["Fan_Vote_Percent"].astype(float).to_numpy()
        real = g["Real_Fan"].astype(float).to_numpy()
        total_real = float(np.sum(real))

        best_beta = 1.0
        best_lambda = 1.0
        best_wmape = float("inf")
        if g.shape[0] >= 3 and float(np.sum(p)) > 0:
            for beta in beta_grid:
                p_adj = np.power(p, float(beta))
                s = float(np.sum(p_adj))
                if s <= 0:
                    continue
                p_adj = p_adj / s
                base_pred = p_adj * total_real
                for lam in lambda_grid:
                    pred = (1.0 - float(lam)) * real + float(lam) * base_pred
                    wmape = float(np.sum(np.abs(pred - real)) / total_real) if total_real > 0 else float("inf")
                    if wmape < best_wmape:
                        best_wmape = wmape
                        best_beta = float(beta)
                        best_lambda = float(lam)

        p_adj = np.power(p, float(best_beta))
        p_adj = p_adj / float(np.sum(p_adj)) if float(np.sum(p_adj)) > 0 else p
        base_pred_votes = p_adj * total_real
        pred_votes = (1.0 - float(best_lambda)) * real + float(best_lambda) * base_pred_votes

        g["Beta"] = best_beta
        g["Lambda"] = best_lambda
        g["Total_Fan_Votes_hat"] = total_real
        g["Pred_Fan_Votes"] = pd.Series(pred_votes, index=g.index).round().astype("Int64")
        g["Pred_Fan_Votes"] = g["Pred_Fan_Votes"].fillna(pd.NA)
        g["Abs_Error"] = (g["Pred_Fan_Votes"].astype("Int64") - g["Real_Fan"].astype("Int64")).abs()
        g["Signed_Error"] = g["Pred_Fan_Votes"].astype("Int64") - g["Real_Fan"].astype("Int64")

        real_sum = float(g["Real_Fan"].sum())
        abs_sum = float(g["Abs_Error"].sum())
        wmape = abs_sum / real_sum if real_sum > 0 else float("nan")

        real = g["Real_Fan"].astype(float).to_numpy()
        pred = g["Pred_Fan_Votes"].astype(float).to_numpy()
        mask = np.isfinite(real) & np.isfinite(pred) & (real >= 0) & (pred >= 0)
        if mask.any():
            rmsle = math.sqrt(float(np.mean((np.log1p(pred[mask]) - np.log1p(real[mask])) ** 2)))
        else:
            rmsle = float("nan")

        if g.shape[0] >= 3:
            r1 = pd.Series(real).rank(method="average").to_numpy()
            r2 = pd.Series(pred).rank(method="average").to_numpy()
            sr = float(np.corrcoef(r1, r2)[0, 1])
        else:
            sr = float("nan")

        week_rows.append(
            {
                "Season": int(season),
                "Week": int(week),
                "n_matched": int(g.shape[0]),
                "Total_Fan_Votes_hat": float(total_real),
                "Beta": float(best_beta),
                "Lambda": float(best_lambda),
                "wMAPE": float(wmape),
                "RMSLE": float(rmsle),
                "Spearman": float(sr),
                "Real_Fan_sum": float(real_sum),
            }
        )

        pred_rows.append(g)

    by_person_week = pd.concat(pred_rows, ignore_index=True) if pred_rows else pd.DataFrame()
    by_week = pd.DataFrame(week_rows).sort_values(["Season", "Week"]) if week_rows else pd.DataFrame()

    if not by_person_week.empty:
        real_all = by_person_week["Real_Fan"].astype(float).to_numpy()
        pred_all = by_person_week["Pred_Fan_Votes"].astype(float).to_numpy()
        abs_all = np.abs(pred_all - real_all)
        denom = float(np.sum(real_all))
        global_wmape = float(np.sum(abs_all) / denom) if denom > 0 else float("nan")
        mask = np.isfinite(real_all) & np.isfinite(pred_all) & (real_all >= 0) & (pred_all >= 0)
        global_rmsle = math.sqrt(float(np.mean((np.log1p(pred_all[mask]) - np.log1p(real_all[mask])) ** 2))) if mask.any() else float("nan")
    else:
        global_wmape = float("nan")
        global_rmsle = float("nan")

    metrics_rows.append(
        {
            "scope": "global",
            "wMAPE": global_wmape,
            "RMSLE": global_rmsle,
            "n_rows": int(by_person_week.shape[0]) if not by_person_week.empty else 0,
        }
    )
    metrics_summary = pd.DataFrame(metrics_rows)
    return by_person_week, by_week, metrics_summary


def build_join_diagnostics(pred: pd.DataFrame, real: pd.DataFrame) -> pd.DataFrame:
    pred_keys = pred[["Season", "Name", "name_key"]].drop_duplicates()
    real_keys = real[["Season", "Name", "name_key"]].drop_duplicates()

    left = pred_keys.merge(real_keys[["Season", "name_key"]], on=["Season", "name_key"], how="left", indicator=True)
    miss_pred = left[left["_merge"] == "left_only"].copy()
    miss_pred["missing_in"] = "real_fan"
    miss_pred = miss_pred[["Season", "Name", "name_key", "missing_in"]]

    right = real_keys.merge(pred_keys[["Season", "name_key"]], on=["Season", "name_key"], how="left", indicator=True)
    miss_real = right[right["_merge"] == "left_only"].copy()
    miss_real["missing_in"] = "predictions"
    miss_real = miss_real[["Season", "Name", "name_key", "missing_in"]]

    out = pd.concat([miss_pred, miss_real], ignore_index=True)
    out = out.sort_values(["missing_in", "Season", "Name"])
    return out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--real_fan", type=str, default=r"d:\MCM_2026_O\data\raw\real_fan.csv")
    parser.add_argument("--pred", type=str, default=r"d:\MCM_2026_O\data\processed\model1_fan_vote_predictions.csv")
    parser.add_argument("--results_dir", type=str, default=r"d:\MCM_2026_O\results")
    args = parser.parse_args()

    real_fan_path = Path(args.real_fan)
    pred_path = Path(args.pred)
    results_dir = Path(args.results_dir)
    results_dir.mkdir(parents=True, exist_ok=True)

    real = load_real_fan_table(real_fan_path)
    pred = load_predictions(pred_path)

    diag = build_join_diagnostics(pred, real)

    merged = pred.merge(
        real[["Season", "name_key", "Name", "Real_Fan_Wan", "Real_Fan", "Evidence"]].rename(columns={"Name": "Name_real"}),
        on=["Season", "name_key"],
        how="inner",
    )
    merged = merged.rename(columns={"Name": "Name_pred"})
    merged["Name"] = merged["Name_pred"]
    merged = merged.drop(columns=["Name_pred"]).copy()

    by_person_week, by_week, metrics = estimate_totals_and_votes(merged)

    out1 = results_dir / "fan_vote_estimates_by_person_week.csv"
    out2 = results_dir / "fan_vote_total_by_week.csv"
    out3 = results_dir / "fan_vote_metrics_summary.csv"
    out4 = results_dir / "fan_vote_join_diagnostics.csv"

    by_person_week.to_csv(out1, index=False, encoding="utf-8-sig")
    by_week.to_csv(out2, index=False, encoding="utf-8-sig")
    metrics.to_csv(out3, index=False, encoding="utf-8-sig")
    diag.to_csv(out4, index=False, encoding="utf-8-sig")

    print(f"Wrote: {out1}")
    print(f"Wrote: {out2}")
    print(f"Wrote: {out3}")
    print(f"Wrote: {out4}")
    print(f"Matched rows: {by_person_week.shape if isinstance(by_person_week, pd.DataFrame) else None}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


"""
最终生成结果：
fan_vote_estimates_by_person_week.csv
    Name_real 姓名
    Real_Fan_Wan 真实粉丝万
    Real_Fan 真实粉丝
    Pred_Fan_Votes 预测粉丝
    implied_total 用该选手当锚点反推总粉丝量级
    Total_Fan_Votes_hat 最终确定总粉丝预测数
    Abs_Error 绝对误差
    Signed_Error 有符号误差
fan_vote_total_by_week.csv
    Season 赛季
    Week 周
    Total_Fan_Votes 总粉丝预测数
    

fan_vote_metrics_summary.csv



"""
