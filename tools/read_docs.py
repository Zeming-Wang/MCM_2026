import argparse
import os
from pathlib import Path


def _iter_files(root: Path, exts: set[str]) -> list[Path]:
    exclude_dir_names = {
        ".git",
        ".conda",
        ".venv",
        "__pycache__",
        ".ipynb_checkpoints",
    }
    matched: list[Path] = []
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in exclude_dir_names]
        for name in filenames:
            p = Path(dirpath) / name
            if p.suffix.lower() in exts:
                matched.append(p)
    return sorted(matched)


def _read_pdf_summary(path: Path) -> dict:
    candidates: list[tuple[str, object]] = []
    try:
        from pypdf import PdfReader as _PdfReader  # type: ignore

        candidates.append(("pypdf", _PdfReader))
    except Exception:
        pass

    try:
        from PyPDF2 import PdfReader as _PdfReader2  # type: ignore

        candidates.append(("PyPDF2", _PdfReader2))
    except Exception:
        pass

    if not candidates:
        raise RuntimeError("Missing PDF reader library: install pypdf or PyPDF2")

    errors: list[str] = []
    for name, ReaderCls in candidates:
        try:
            reader = ReaderCls(str(path))
            pages = list(reader.pages)
            extracted: list[str] = []
            for page in pages:
                try:
                    extracted.append(page.extract_text() or "")
                except Exception:
                    extracted.append("")
            text = "\n".join(extracted)
            sample = text.strip().replace("\r", " ").replace("\n", " ")
            return {
                "kind": "pdf",
                "backend": name,
                "pages": len(pages),
                "extracted_chars": len(text),
                "sample": sample[:240],
                "text_is_empty": len(sample) == 0,
            }
        except Exception as e:
            errors.append(f"{name}: {e}")
            continue

    raise RuntimeError("PDF read failed: " + " | ".join(errors))


def _read_csv_summary(path: Path) -> dict:
    try:
        import pandas as pd  # type: ignore
    except Exception:
        import csv

        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            header = next(reader, [])
            rows = sum(1 for _ in reader)
        return {
            "kind": "csv",
            "rows": rows,
            "cols": len(header),
            "columns": header,
        }

    encodings = ["utf-8-sig", "utf-8", "gb18030", "gbk", "latin1"]
    last_err = None
    for enc in encodings:
        try:
            df = pd.read_csv(path, encoding=enc)
            return {
                "kind": "csv",
                "rows": int(df.shape[0]),
                "cols": int(df.shape[1]),
                "columns": [str(c) for c in df.columns.tolist()],
                "encoding": enc,
            }
        except Exception as e:
            last_err = e
            continue
    raise RuntimeError(f"CSV decode failed: {last_err}")


def _read_excel_summary(path: Path) -> dict:
    try:
        import pandas as pd  # type: ignore

        xls = pd.ExcelFile(path)
        sheet_names = list(xls.sheet_names)
        df0 = pd.read_excel(xls, sheet_name=sheet_names[0]) if sheet_names else None
        return {
            "kind": "excel",
            "sheets": sheet_names,
            "first_sheet_shape": None if df0 is None else [int(df0.shape[0]), int(df0.shape[1])],
            "first_sheet_columns": None if df0 is None else [str(c) for c in df0.columns.tolist()],
        }
    except Exception:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        if sheets:
            ws = wb[sheets[0]]
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            first_row = [] if first_row is None else [None if v is None else str(v) for v in first_row]
        else:
            first_row = []
        return {"kind": "excel", "sheets": sheets, "first_row": first_row}


def _summarize_file(path: Path) -> dict:
    suf = path.suffix.lower()
    if suf == ".pdf":
        return _read_pdf_summary(path)
    if suf == ".csv":
        return _read_csv_summary(path)
    if suf in {".xlsx", ".xlsm"}:
        return _read_excel_summary(path)
    raise ValueError(f"Unsupported file type: {path}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=str, default=str(Path(__file__).resolve().parents[1]))
    parser.add_argument("--preview", type=int, default=0)
    args = parser.parse_args()

    root = Path(args.root).resolve()
    exts = {".pdf", ".csv", ".xlsx", ".xlsm"}
    files = _iter_files(root, exts=exts)
    if not files:
        print(f"No matching files found under: {root}")
        return 0

    failures: list[tuple[Path, str]] = []
    for p in files:
        rel = str(p.relative_to(root)) if p.is_relative_to(root) else str(p)
        try:
            info = _summarize_file(p)
            line = f"OK  {rel}  {info}"
            if info.get("kind") == "pdf" and info.get("text_is_empty"):
                line = f"WARN {rel}  {info}"
            print(line)
            if args.preview and info.get("kind") == "pdf":
                print(info.get("sample", "")[: int(args.preview)])
        except Exception as e:
            failures.append((p, str(e)))
            print(f"FAIL {rel}  {e}")

    if failures:
        print(f"\nFAILED ({len(failures)}/{len(files)})")
        for p, err in failures[:50]:
            rel = str(p.relative_to(root)) if p.is_relative_to(root) else str(p)
            print(f"- {rel}: {err}")
        return 2

    print(f"\nSUCCESS ({len(files)}/{len(files)})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
